"""Aligner Window — Abbyy Aligner style. Standalone alignment tool."""

from pathlib import Path

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QListWidget, QListWidgetItem, QLabel, QPushButton, QComboBox,
    QFileDialog, QMessageBox, QProgressBar, QGroupBox, QAbstractItemView, QMenu,
    QLineEdit, QDialog, QDialogButtonBox,
)
from PySide6.QtCore import Qt, Signal, QPoint, QRect, QThread
from PySide6.QtGui import QPainter, QPen, QColor, QFont, QPainterPath, QBrush, QAction

from ruzh_translator.models.base import get_session
from ruzh_translator.models.segment import AlignmentPair, Segment
from ruzh_translator.services.alignment_service import align_documents, save_alignment_to_db
from ruzh_translator.services.import_service import import_file, import_and_segment
from ruzh_translator.services.project_service import create_project, get_project, update_project_status
from ruzh_translator.config import IMPORT_FORMATS, SOURCE_LANGS, TARGET_LANGS
from ruzh_translator.utils.text_utils import split_sentences


class ConnectorWidget(QWidget):
    """Draws bezier connector lines between aligned source/target items."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._pairs = []
        self._selected = -1
        self._src_n = 0
        self._tgt_n = 0
        self.setFixedWidth(56)
        self.setMinimumHeight(100)

    def set_data(self, pairs, src_n, tgt_n):
        self._pairs = pairs
        self._src_n = max(src_n, 1)
        self._tgt_n = max(tgt_n, 1)
        self._selected = -1
        self.update()

    def select(self, idx):
        self._selected = idx
        self.update()

    def paintEvent(self, e):
        if not self._pairs:
            return
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        h = self.height()
        sh = h / self._src_n
        th = h / self._tgt_n
        for i, (si, ti, conf) in enumerate(self._pairs):
            if conf < 0:
                continue
            pen = QPen(QColor("#FF5722" if i == self._selected else ("#4CAF50" if conf > 0.8 else "#FF9800" if conf > 0.5 else "#F44336")), 3 if i == self._selected else 2)
            p.setPen(pen)
            sy, ty = int(sh * (si + 0.5)), int(th * (ti + 0.5))
            path = QPainterPath(); path.moveTo(4, sy); path.cubicTo(22, sy, 34, ty, 52, ty)
            p.drawPath(path)
            mx, my = 28, (sy + ty) // 2
            p.setBrush(QBrush(pen.color())); p.setPen(Qt.NoPen)
            p.drawEllipse(QPoint(mx, my), 8, 8)
            p.setPen(Qt.white); p.setFont(QFont("sans-serif", 7))
            p.drawText(QRect(mx - 8, my - 6, 16, 12), Qt.AlignCenter, str(int(conf * 100)))
        p.end()


class AlignWorker(QThread):
    """Paragraph-by-paragraph alignment with fallback to sequential."""
    finished = Signal(list)
    progress = Signal(int, int)
    status_msg = Signal(str)
    error = Signal(str)

    def __init__(self, src_paras, tgt_paras, use_semantic=True):
        super().__init__()
        self.src_paras = src_paras
        self.tgt_paras = tgt_paras
        self.use_semantic = use_semantic

    def run(self):
        try:
            # Try loading model
            model = None
            if self.use_semantic:
                try:
                    from sentence_transformers import SentenceTransformer
                    self.status_msg.emit("正在加载 LaBSE 模型...")
                    model = SentenceTransformer("LaBSE")
                    self.status_msg.emit("模型加载完成，开始对齐...")
                except Exception as e:
                    self.status_msg.emit(f"模型加载失败({e})，使用顺序对齐...")
                    model = None

            all_pairs = []
            pair_idx = 0
            total = max(len(self.src_paras), len(self.tgt_paras))

            for pi in range(total):
                src_para = self.src_paras[pi] if pi < len(self.src_paras) else ""
                tgt_para = self.tgt_paras[pi] if pi < len(self.tgt_paras) else ""

                src_sents = split_sentences(src_para, "ru")
                tgt_sents = split_sentences(tgt_para, "zh-CN")

                if not src_sents or not tgt_sents:
                    self.progress.emit(pi + 1, total)
                    continue

                if model is not None:
                    # Semantic matching
                    import numpy as np
                    src_emb = model.encode(src_sents, show_progress_bar=False)
                    tgt_emb = model.encode(tgt_sents, show_progress_bar=False)
                    src_n = src_emb / np.linalg.norm(src_emb, axis=1, keepdims=True)
                    tgt_n = tgt_emb / np.linalg.norm(tgt_emb, axis=1, keepdims=True)

                    if len(src_sents) >= len(tgt_sents):
                        for si, srow in enumerate(src_n):
                            scores = np.dot(tgt_n, srow)
                            best_ti = int(np.argmax(scores))
                            all_pairs.append((pair_idx + si, pair_idx + best_ti, float(scores[best_ti])))
                    else:
                        for ti, trow in enumerate(tgt_n):
                            scores = np.dot(src_n, trow)
                            best_si = int(np.argmax(scores))
                            all_pairs.append((pair_idx + best_si, pair_idx + ti, float(scores[best_si])))
                else:
                    # Sequential fallback: simple zip
                    for k in range(max(len(src_sents), len(tgt_sents))):
                        all_pairs.append((pair_idx + k, pair_idx + k, 0.5))

                pair_idx += max(len(src_sents), len(tgt_sents))
                self.progress.emit(pi + 1, total)

            self.finished.emit(all_pairs)
        except Exception as e:
            self.error.emit(str(e))


class AlignerWindow(QMainWindow):
    """Standalone Abbyy-style aligner."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("🔗 语料对齐 — Aligner")
        self.resize(1400, 900)
        self._session = get_session()
        self._project_id = None
        self._src_sents = []
        self._tgt_sents = []
        self._pairs_data = []
        self._selected_pair = -1
        self._setup_ui()

    def _setup_ui(self):
        c = QWidget()
        self.setCentralWidget(c)
        layout = QVBoxLayout(c)
        layout.setContentsMargins(8, 8, 8, 8)

        # Top bar
        top = QHBoxLayout()
        top.addWidget(QLabel("<b>🔗 语料对齐</b>"))

        self._proj_combo = QComboBox()
        self._proj_combo.addItem("独立对齐（不关联项目）", None)
        self._refresh_projects()
        self._proj_combo.currentIndexChanged.connect(self._on_proj_changed)
        top.addWidget(self._proj_combo, 1)
        top.addWidget(QPushButton("＋新建", clicked=self._on_new_project))

        top.addWidget(QLabel("源语:"))
        top.addWidget(QPushButton("📂 导入", clicked=self._on_import_src))
        top.addWidget(QLabel("目标语:"))
        top.addWidget(QPushButton("📂 导入", clicked=self._on_import_tgt))

        self._method_combo = QComboBox()
        self._method_combo.addItems(["语义 (LaBSE)", "顺序"])
        top.addWidget(self._method_combo)

        align_btn = QPushButton("🔗 对齐")
        align_btn.clicked.connect(self._on_align)
        align_btn.setStyleSheet("background:#FF9800;color:white;padding:6px 16px;")
        top.addWidget(align_btn)
        layout.addLayout(top)

        self._progress = QProgressBar()
        self._progress.setVisible(False)
        layout.addWidget(self._progress)

        # Body: Source | Connector | Target
        body = QHBoxLayout()
        body.setSpacing(0)

        sg = QGroupBox("源语（俄语）")
        sl = QVBoxLayout(sg)
        self._src_list = QListWidget()
        self._src_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._src_list.setStyleSheet("QListWidget::item{padding:8px;border-bottom:1px solid #EEE;font-size:13px;} QListWidget::item:selected{background:#E3F2FD;}")
        self._src_list.currentRowChanged.connect(self._on_src_sel)
        sl.addWidget(self._src_list)
        body.addWidget(sg, 4)

        self._connector = ConnectorWidget()
        body.addWidget(self._connector, 0)

        tg = QGroupBox("目标语（中文）")
        tl = QVBoxLayout(tg)
        self._tgt_list = QListWidget()
        self._tgt_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self._tgt_list.setStyleSheet("QListWidget::item{padding:8px;border-bottom:1px solid #EEE;font-size:13px;} QListWidget::item:selected{background:#FFF9C4;}")
        self._tgt_list.currentRowChanged.connect(self._on_tgt_sel)
        tl.addWidget(self._tgt_list)
        body.addWidget(tg, 4)

        self._src_list.verticalScrollBar().valueChanged.connect(self._tgt_list.verticalScrollBar().setValue)
        self._tgt_list.verticalScrollBar().valueChanged.connect(self._src_list.verticalScrollBar().setValue)
        layout.addLayout(body, 1)

        # Bottom
        bot = QHBoxLayout()
        for text, fn in [("✂️拆分", self._on_split), ("🔗合并", self._on_merge), ("⚠️错配", self._on_misalign), ("🔗重连", self._on_relink)]:
            bot.addWidget(QPushButton(text, clicked=fn))
        bot.addStretch()
        save_btn = QPushButton("💾 保存到项目"); save_btn.clicked.connect(self._on_save)
        save_btn.setStyleSheet("background:#1565C0;color:white;padding:8px 16px;")
        bot.addWidget(save_btn)
        bot.addWidget(QPushButton("📤 导出 XLSX", clicked=self._on_export))
        layout.addLayout(bot)

        self._status = QLabel("📂 请导入源语和目标语文件，然后点击「对齐」")
        self._status.setStyleSheet("color:#9E9E9E;padding:4px;")
        layout.addWidget(self._status)

    def _refresh_projects(self):
        from ruzh_translator.services.project_service import list_projects
        # Keep first item, re-add projects
        while self._proj_combo.count() > 1:
            self._proj_combo.removeItem(1)
        for p in list_projects(self._session):
            self._proj_combo.addItem(f"{p.name}", p.id)

    def _on_new_project(self):
        from PySide6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "新建项目", "项目名称:")
        if ok and name.strip():
            from ruzh_translator.services.project_service import create_project
            create_project(self._session, name.strip())
            self._refresh_projects()

    def _on_proj_changed(self):
        pid = self._proj_combo.currentData()
        if pid:
            self._project_id = pid
            segs = self._session.query(Segment).filter(Segment.project_id == pid).order_by(Segment.paragraph_index, Segment.segment_index).all()
            src, tgt = [], []
            for s in segs:
                t = (s.source_text or "").strip()
                if not t: continue
                if any(0x0400 <= ord(c) <= 0x04FF for c in t[:30]):
                    src.append(t)
                elif s.target_text and s.target_text.strip():
                    src.append(t); tgt.append(s.target_text.strip())
                else:
                    tgt.append(t)
            if src:
                self._src_sents = src; self._tgt_sents = tgt; self._populate_lists()
                self._status.setText(f"已加载项目: {len(src)} 源语句 | {len(tgt)} 目标语句")

    def _on_import_src(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入源语文件", "", ";;".join(f"{d} (*{e})" for e, d in IMPORT_FORMATS.items()))
        if not path: return
        data = import_file(path)
        self._src_paras = data["paragraphs"]  # Keep paragraphs for alignment
        # Also flatten for display
        self._src_sents = []
        for para in self._src_paras:
            self._src_sents.extend(split_sentences(para, data["language"]))
        self._src_sents = [s.strip() for s in self._src_sents if s.strip()]
        self._populate_lists()
        self._status.setText(f"源语: {len(self._src_sents)} 句 ({len(self._src_paras)} 段) — 请导入目标语文件")
        if self._tgt_sents:
            self._on_align()  # Auto-align if both files loaded

    def _on_import_tgt(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入目标语文件", "", ";;".join(f"{d} (*{e})" for e, d in IMPORT_FORMATS.items()))
        if not path: return
        data = import_file(path)
        self._tgt_paras = data["paragraphs"]
        self._tgt_sents = []
        for para in self._tgt_paras:
            self._tgt_sents.extend(split_sentences(para, "zh-CN"))
        self._tgt_sents = [s.strip() for s in self._tgt_sents if s.strip()]
        self._populate_lists()
        self._status.setText(f"源语: {len(self._src_sents)} 句 | 目标语: {len(self._tgt_sents)} 句 — 自动对齐中...")
        if self._src_sents:
            self._on_align()  # Auto-align

    def _populate_lists(self):
        self._src_list.clear(); self._tgt_list.clear()
        for s in self._src_sents: self._src_list.addItem(QListWidgetItem(s[:100]))
        for s in self._tgt_sents: self._tgt_list.addItem(QListWidgetItem(s[:100]))

    def _on_align(self):
        if not self._src_sents or not self._tgt_sents:
            QMessageBox.warning(self, "提示", "请先导入源语和目标语文件"); return

        # Use paragraphs if available, otherwise wrap sentences as single paragraphs
        src_paras = getattr(self, '_src_paras', None) or self._src_sents
        tgt_paras = getattr(self, '_tgt_paras', None) or self._tgt_sents

        total = max(len(src_paras), len(tgt_paras))
        self._progress.setVisible(True)
        self._progress.setRange(0, total)
        self._progress.setValue(0)
        self._status.setText(f"⏳ 对齐中... 0/{total} 段")

        use_semantic = "语义" in self._method_combo.currentText()
        self._worker = AlignWorker(src_paras, tgt_paras, use_semantic=use_semantic)
        self._worker.status_msg.connect(lambda msg: self._status.setText(f"⏳ {msg}"))
        self._worker.progress.connect(lambda cur, tot: (
            self._progress.setValue(cur),
            self._status.setText(f"⏳ 对齐中... {cur}/{tot} 段 ({cur*100//tot}%)")
        ))
        self._worker.finished.connect(self._on_align_done)
        self._worker.error.connect(self._on_align_error)
        self._worker.start()

    def _on_align_done(self, data):
        self._pairs_data = data
        self._connector.set_data(self._pairs_data, len(self._src_sents), len(self._tgt_sents))
        self._apply_colors()
        hi = sum(1 for _, _, c in self._pairs_data if c > 0.8)
        self._status.setText(
            f"✓ 对齐完成: {len(self._pairs_data)} 对 | "
            f"🟢{hi} 🟠{sum(1 for _,_,c in self._pairs_data if 0.5<=c<=0.8)} "
            f"🔴{sum(1 for _,_,c in self._pairs_data if c<0.5)}"
        )
        self._progress.setVisible(False)

    def _on_align_error(self, msg):
        self._progress.setVisible(False)
        QMessageBox.critical(self, "对齐失败", msg)

    def _apply_colors(self):
        # Reset all items first
        for i in range(self._src_list.count()):
            item = self._src_list.item(i)
            item.setBackground(QColor("#FFFFFF"))
            item.setForeground(QColor("#212121"))
        for i in range(self._tgt_list.count()):
            item = self._tgt_list.item(i)
            item.setBackground(QColor("#FFFFFF"))
            item.setForeground(QColor("#212121"))

        # Color only aligned pairs
        aligned_src = set(); aligned_tgt = set()
        for si, ti, conf in self._pairs_data:
            aligned_src.add(si); aligned_tgt.add(ti)
            c = QColor("#C8E6C9" if conf > 0.8 else "#FFE0B2" if conf > 0.5 else "#FFCDD2")
            if si < self._src_list.count():
                self._src_list.item(si).setBackground(c)
            if ti < self._tgt_list.count():
                self._tgt_list.item(ti).setBackground(c)

        # Only gray unaligned if there ARE pairs
        if self._pairs_data:
            for i in range(self._src_list.count()):
                if i not in aligned_src:
                    self._src_list.item(i).setForeground(QColor("#BDBDBD"))
            for i in range(self._tgt_list.count()):
                if i not in aligned_tgt:
                    self._tgt_list.item(i).setForeground(QColor("#BDBDBD"))

    def _on_src_sel(self, r):
        for i, (si, ti, _) in enumerate(self._pairs_data):
            if si == r: self._selected_pair = i; self._connector.select(i); self._tgt_list.blockSignals(True); self._tgt_list.setCurrentRow(ti); self._tgt_list.blockSignals(False); return

    def _on_tgt_sel(self, r):
        for i, (si, ti, _) in enumerate(self._pairs_data):
            if ti == r: self._selected_pair = i; self._connector.select(i); self._src_list.blockSignals(True); self._src_list.setCurrentRow(si); self._src_list.blockSignals(False); return

    def _on_split(self):
        r = self._src_list.currentRow()
        if r < 0: QMessageBox.warning(self, "提示", "请先选中要拆分的源语句子"); return
        t = self._src_sents[r]
        # Smart split: find best break point
        mid = len(t) // 2
        for punc in [',', '，', ';', '；', '—', ' –']:
            idx = t.rfind(punc, 0, mid + len(t) // 4)
            if idx > 0: mid = idx + 1; break

        # Show split preview dialog
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QDialogButtonBox
        dlg = QDialog(self); dlg.setWindowTitle("拆分句子 — 确认拆分位置"); dlg.setMinimumWidth(550)
        l = QVBoxLayout(dlg)
        l.addWidget(QLabel(f"<b>原句 (行 {r+1}):</b>"))
        l.addWidget(QLabel(t))
        l.addWidget(QLabel(f"<b>拆分位置:</b> (第 {mid} 个字符处)"))
        part1 = QLineEdit(t[:mid].strip()); part2 = QLineEdit(t[mid:].strip())
        l.addWidget(QLabel("上半句:")); l.addWidget(part1)
        l.addWidget(QLabel("下半句:")); l.addWidget(part2)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        l.addWidget(btns)
        if not dlg.exec(): return

        p1, p2 = part1.text().strip(), part2.text().strip()
        if not p1 or not p2: return
        self._src_sents[r] = p1; self._src_sents.insert(r + 1, p2)
        new = []
        for si, ti, conf in self._pairs_data:
            if si == r: new.append((si, ti, conf)); new.append((si + 1, -1, 0.0))
            elif si > r: new.append((si + 1, ti, conf))
            else: new.append((si, ti, conf))
        self._pairs_data = new
        self._populate_lists(); self._apply_colors()
        self._connector.set_data(self._pairs_data, len(self._src_sents), len(self._tgt_sents))

    def _on_merge(self):
        # Merge selected with previous — or merge if two rows are selected
        rows = sorted(set(i.row() for i in self._src_list.selectedIndexes()))
        if len(rows) >= 2:
            # Merge two selected rows
            r1, r2 = rows[0], rows[1]
            self._src_sents[r1] += " " + self._src_sents.pop(r2)
            new = []
            for si, ti, conf in self._pairs_data:
                if si == r2: continue  # drop
                elif si > r2: new.append((si - 1, ti, conf))
                else: new.append((si, ti, conf))
            self._pairs_data = new
        else:
            r = self._src_list.currentRow()
            if r <= 0: QMessageBox.warning(self, "提示", "请选中要合并的两行（Ctrl+点击），或选中第二行合并到上一行"); return
            self._src_sents[r - 1] += " " + self._src_sents.pop(r)
            new = [(si - 1 if si > r else si, ti, conf) for si, ti, conf in self._pairs_data if si != r]
            self._pairs_data = new
        self._populate_lists(); self._apply_colors()
        self._connector.set_data(self._pairs_data, len(self._src_sents), len(self._tgt_sents))

    def _on_misalign(self):
        if self._selected_pair < 0: return
        si, ti, _ = self._pairs_data[self._selected_pair]
        self._pairs_data[self._selected_pair] = (si, ti, 0.0)
        self._apply_colors(); self._connector.set_data(self._pairs_data, len(self._src_sents), len(self._tgt_sents))

    def _on_relink(self):
        sr, tr = self._src_list.currentRow(), self._tgt_list.currentRow()
        if sr < 0 or tr < 0: return
        for i, (si, ti, _) in enumerate(self._pairs_data):
            if si == sr or ti == tr: self._pairs_data[i] = (sr, tr, 0.5); self._connector.set_data(self._pairs_data, len(self._src_sents), len(self._tgt_sents)); self._apply_colors(); return
        self._pairs_data.append((sr, tr, 0.5)); self._connector.set_data(self._pairs_data, len(self._src_sents), len(self._tgt_sents)); self._apply_colors()

    def _on_save(self):
        if not self._pairs_data: QMessageBox.warning(self, "提示", "没有可保存的对齐数据"); return
        if not self._project_id:
            p = create_project(self._session, "对齐项目"); self._project_id = p.id; self._proj_combo.addItem(p.name, p.id)
        self._session.query(AlignmentPair).filter(AlignmentPair.project_id == self._project_id).delete()
        pairs = [{"para_index":0,"sent_index":i,"source_text":self._src_sents[si] if si<len(self._src_sents) else "", "target_text":self._tgt_sents[ti] if ti<len(self._tgt_sents) else "", "confidence":conf, "is_manually_corrected":conf<0.8} for i,(si,ti,conf) in enumerate(self._pairs_data)]
        save_alignment_to_db(self._session, self._project_id, "", pairs)
        update_project_status(self._session, self._project_id, "alignment")
        QMessageBox.information(self, "保存成功", f"已保存 {len(pairs)} 个对齐对")

    def _on_export(self):
        if not self._pairs_data: return
        path, _ = QFileDialog.getSaveFileName(self, "导出对齐结果", "aligned.xlsx", "Excel (*.xlsx)")
        if not path: return
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "对齐结果"
        for c, h in enumerate(["序号", "源语", "目标语", "置信度"], 1):
            ws.cell(1, c, h)
        for i, (si, ti, conf) in enumerate(self._pairs_data):
            ws.cell(i+2, 1, i+1)
            ws.cell(i+2, 2, self._src_sents[si] if si<len(self._src_sents) else "")
            ws.cell(i+2, 3, self._tgt_sents[ti] if ti<len(self._tgt_sents) else "")
            ws.cell(i+2, 4, round(conf, 4))
        wb.save(path)
        QMessageBox.information(self, "导出成功", f"已导出到: {path}")
