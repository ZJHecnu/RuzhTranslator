"""Concordance Window — 3-tab: Raw Doc | KWIC+Tag | Statistics."""

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QLineEdit, QPushButton, QLabel, QListWidget, QListWidgetItem,
    QTextEdit, QFileDialog, QMessageBox, QGroupBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QComboBox, QTabWidget,
    QDialog, QDialogButtonBox,
)
from PySide6.QtCore import Qt

from ruzh_translator.models.base import get_session
from ruzh_translator.models.tm import TranslationMemoryEntry
from ruzh_translator.services.tm_service import concordance_search, add_tm_entry
from ruzh_translator.models.segment import AlignmentPair
from ruzh_translator.services.project_service import list_projects


class ConcordanceWindow(QMainWindow):
    """ParaConc-style: Raw Document → KWIC Search & Tag → Statistics."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("📊 平行语料 — Concordance")
        self.resize(1200, 850)
        self._session = get_session()
        self._raw_src = ""
        self._raw_tgt = ""
        self._kwic_tags = {}
        self._setup_ui()
        self._refresh_tm()

    def _setup_ui(self):
        c = QWidget(); self.setCentralWidget(c)
        layout = QVBoxLayout(c); layout.setContentsMargins(8, 8, 8, 8)

        # Toolbar
        tools = QHBoxLayout()
        tools.addWidget(QLabel("<b>📊 平行语料库</b>"))

        self._proj_combo = QComboBox()
        self._proj_combo.addItem("从项目加载语料...", None)
        for p in list_projects(self._session):
            self._proj_combo.addItem(f"📋 {p.name}", p.id)
        self._proj_combo.currentIndexChanged.connect(self._on_load_project)
        tools.addWidget(self._proj_combo, 1)

        tools.addWidget(QPushButton("📂 TMX", clicked=self._on_import_tmx))
        tools.addWidget(QPushButton("📂 XLSX", clicked=self._on_import_xlsx))
        layout.addLayout(tools)

        # Three tabs
        tabs = QTabWidget()

        # Tab 1: Raw Document
        tabs.addTab(self._make_raw_tab(), "📄 原始文档")

        # Tab 2: KWIC Search + Tag
        tabs.addTab(self._make_kwic_tab(), "🔍 KWIC检索与标注")

        # Tab 3: Statistics
        tabs.addTab(self._make_stats_tab(), "📈 统计结果")

        layout.addWidget(tabs, 1)

        # Bottom TM browser
        tm_g = QGroupBox("翻译记忆库")
        tm_l = QVBoxLayout(tm_g)
        tm_h = QHBoxLayout()
        tm_h.addWidget(QLabel("共 0 条"))
        tm_h.addStretch()
        tm_h.addWidget(QPushButton("🔄 刷新", clicked=self._refresh_tm))
        tm_l.addLayout(tm_h)
        self._tm_browse = QListWidget()
        self._tm_browse.setMaximumHeight(120)
        self._tm_browse.itemDoubleClicked.connect(self._on_tm_click)
        tm_l.addWidget(self._tm_browse)
        layout.addWidget(tm_g)

    # ── Tab 1: Raw Document ──────────────────────────────────

    def _make_raw_tab(self):
        w = QWidget()
        split = QSplitter(Qt.Horizontal)
        sg = QGroupBox("源语文档"); sl = QVBoxLayout(sg)
        self._raw_src_view = QTextEdit(); self._raw_src_view.setReadOnly(True)
        sl.addWidget(self._raw_src_view); split.addWidget(sg)
        tg = QGroupBox("目标语文档"); tl = QVBoxLayout(tg)
        self._raw_tgt_view = QTextEdit(); self._raw_tgt_view.setReadOnly(True)
        tl.addWidget(self._raw_tgt_view); split.addWidget(tg)
        layout = QVBoxLayout(w); layout.addWidget(split)
        return w

    # ── Tab 2: KWIC + Tag ────────────────────────────────────

    def _make_kwic_tab(self):
        w = QWidget()
        layout = QVBoxLayout(w)

        # Search bar
        bar = QHBoxLayout()
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("输入关键词检索上下文 (KWIC)...")
        self._search_edit.returnPressed.connect(self._on_search)
        bar.addWidget(self._search_edit, 1)
        bar.addWidget(QPushButton("🔍 检索", clicked=self._on_search))

        self._tag_combo = QComboBox()
        self._tag_combo.setEditable(True)
        self._tag_combo.setPlaceholderText("选择或输入标签...")
        self._tag_combo.setMinimumWidth(100)
        bar.addWidget(self._tag_combo)
        bar.addWidget(QPushButton("🏷️ 标签管理", clicked=self._on_manage_tags))
        bar.addWidget(QPushButton("📤 导出KWIC", clicked=self._on_export))
        bar.addWidget(QPushButton("📈 查看统计", clicked=lambda: self._tab_widget().setCurrentIndex(2)))
        layout.addLayout(bar)

        # KWIC table
        self._kwic_table = QTableWidget()
        self._kwic_table.setColumnCount(5)
        self._kwic_table.setHorizontalHeaderLabels(["#", "左上下文", "关键词", "右上下文", "标签"])
        self._kwic_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self._kwic_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self._kwic_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Fixed)
        self._kwic_table.setColumnWidth(4, 110)
        self._kwic_table.setAlternatingRowColors(True)
        self._kwic_table.setSelectionBehavior(QTableWidget.SelectRows)
        self._kwic_table.cellDoubleClicked.connect(self._on_kwic_click)
        layout.addWidget(self._kwic_table, 1)

        # Detail
        det = QHBoxLayout()
        self._detail_src = QTextEdit(); self._detail_src.setReadOnly(True); self._detail_src.setMaximumHeight(80)
        self._detail_tgt = QTextEdit(); self._detail_tgt.setReadOnly(True); self._detail_tgt.setMaximumHeight(80)
        det.addWidget(QLabel("源语:")); det.addWidget(self._detail_src)
        det.addWidget(QLabel("目标语:")); det.addWidget(self._detail_tgt)
        layout.addLayout(det)
        return w

    def _tab_widget(self):
        return self.centralWidget().layout().itemAt(2).widget()

    # ── Tab 3: Statistics ─────────────────────────────────────

    def _make_stats_tab(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        bar = QHBoxLayout()
        bar.addWidget(QLabel("<b>标注统计</b>"))
        bar.addStretch()
        bar.addWidget(QPushButton("🔄 刷新统计", clicked=self._refresh_stats))
        bar.addWidget(QPushButton("📤 导出标注数据", clicked=self._on_export_tagged))
        layout.addLayout(bar)

        self._stats_view = QTextEdit()
        self._stats_view.setReadOnly(True)
        self._stats_view.setStyleSheet("font-family: 'PingFang SC', monospace; font-size: 13px;")
        layout.addWidget(self._stats_view)
        return w

    def _refresh_stats(self):
        """Show tagging statistics."""
        if not self._kwic_tags:
            self._stats_view.setPlainText("暂无标注数据。请先在「KWIC检索与标注」中检索并标注。")
            return

        from collections import Counter
        tags = [t for t in self._kwic_tags.values() if t]
        counter = Counter(tags)
        total = len(self._kwic_tags)
        tagged = len(tags)

        lines = [
            f"=== 标注统计 ===\n",
            f"总索引行: {total}",
            f"已标注:   {tagged} ({tagged/total*100:.1f}%)" if total else "已标注: 0",
            f"未标注:   {total - tagged}\n",
            f"--- 标签分布 ---",
        ]
        for tag, count in counter.most_common():
            bar = "█" * (count * 30 // max(counter.values(), 1))
            lines.append(f"  {tag:15s} {count:4d}  {bar}")

        self._stats_view.setPlainText("\n".join(lines))

    def _on_export_tagged(self):
        """Export tagged KWIC data."""
        path, _ = QFileDialog.getSaveFileName(self, "导出标注数据", "tagged_kwic.xlsx", "Excel (*.xlsx)")
        if not path: return
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title = "标注数据"
        for c, h in enumerate(["#","左上下文","关键词","右上下文","完整源语","目标语","标签"], 1):
            ws.cell(1, c, h)
        row = 2
        data_idx = []
        for r in range(self._kwic_table.rowCount()):
            item0 = self._kwic_table.item(r, 0)
            if not item0: continue
            data = item0.data(Qt.UserRole) if hasattr(item0, 'data') else {}
            ws.cell(row, 1, r + 1)
            ws.cell(row, 2, self._kwic_table.item(r, 1).text() if self._kwic_table.item(r, 1) else "")
            ws.cell(row, 3, self._kwic_table.item(r, 2).text() if self._kwic_table.item(r, 2) else "")
            ws.cell(row, 4, self._kwic_table.item(r, 3).text() if self._kwic_table.item(r, 3) else "")
            ws.cell(row, 5, data.get("source_text", "") if isinstance(data, dict) else "")
            ws.cell(row, 6, data.get("target_text", "") if isinstance(data, dict) else "")
            tag_w = self._kwic_table.cellWidget(r, 4)
            ws.cell(row, 7, tag_w.currentText() if isinstance(tag_w, QComboBox) else "")
            row += 1
        wb.save(path)
        QMessageBox.information(self, "导出成功", f"已导出 {row-2} 条标注数据到:\n{path}")

    # ── KWIC Search & Tag ────────────────────────────────────

    def _on_search(self):
        query = self._search_edit.text().strip()
        if not query: return
        results = concordance_search(self._session, query, limit=200)
        self._kwic_table.setRowCount(len(results))
        self._kwic_tags = {}
        from PySide6.QtGui import QFont, QColor

        for i, r in enumerate(results):
            self._kwic_table.setItem(i, 0, QTableWidgetItem(str(i+1)))
            src = r["source_text"]
            idx = src.lower().find(query.lower())
            left = src[max(0, idx-30):idx] if idx >= 0 else src[:30]
            right = src[idx+len(query):idx+len(query)+30] if idx >= 0 else ""
            keyword = src[idx:idx+len(query)] if idx >= 0 else query
            self._kwic_table.setItem(i, 1, QTableWidgetItem(f"...{left}"))
            kw = QTableWidgetItem(keyword)
            kw.setForeground(QColor("#FF0000"))
            f = QFont(); f.setBold(True); kw.setFont(f)
            self._kwic_table.setItem(i, 2, kw)
            self._kwic_table.setItem(i, 3, QTableWidgetItem(f"{right}..."))
            # Tag dropdown
            cb = QComboBox(); cb.addItem("")
            cb.addItems(self._get_tag_list())
            self._kwic_table.setCellWidget(i, 4, cb)
            self._kwic_table.item(i, 0).setData(Qt.UserRole, r)

    def _get_tag_list(self):
        return [self._tag_combo.itemText(j) for j in range(self._tag_combo.count()) if self._tag_combo.itemText(j)]

    def _on_kwic_click(self, row, _):
        item = self._kwic_table.item(row, 0)
        if item:
            r = item.data(Qt.UserRole)
            if isinstance(r, dict):
                self._detail_src.setPlainText(r.get("source_text", ""))
                self._detail_tgt.setPlainText(r.get("target_text", ""))

    def _on_manage_tags(self):
        dlg = QDialog(self); dlg.setWindowTitle("标签管理"); dlg.setMinimumWidth(400)
        l = QVBoxLayout(dlg)
        l.addWidget(QLabel("设置常用标签（每行一个）:"))
        edit = QTextEdit()
        edit.setPlainText("\n".join(self._get_tag_list()))
        l.addWidget(edit)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        l.addWidget(btns)
        if dlg.exec():
            tags = [t.strip() for t in edit.toPlainText().split("\n") if t.strip()]
            self._tag_combo.clear(); self._tag_combo.addItems(tags)
            if tags: self._tag_combo.setCurrentIndex(0)
            # Refresh all KWIC row combos
            for r in range(self._kwic_table.rowCount()):
                w = self._kwic_table.cellWidget(r, 4)
                if isinstance(w, QComboBox):
                    cur = w.currentText(); w.blockSignals(True)
                    w.clear(); w.addItem(""); w.addItems(tags)
                    w.setCurrentText(cur); w.blockSignals(False)

    # ── Import ───────────────────────────────────────────────

    def _on_load_project(self):
        pid = self._proj_combo.currentData()
        if not pid: return
        pairs = self._session.query(AlignmentPair).filter(AlignmentPair.project_id == pid).all()
        if not pairs:
            QMessageBox.information(self, "提示", "该项目没有对齐数据")
            return
        src_lines = []; tgt_lines = []
        for p in pairs:
            if p.source_text and p.target_text:
                add_tm_entry(self._session, p.source_text, p.target_text, project_id=pid)
                src_lines.append(p.source_text); tgt_lines.append(p.target_text)
        self._raw_src = "\n\n".join(src_lines)
        self._raw_tgt = "\n\n".join(tgt_lines)
        self._raw_src_view.setPlainText(self._raw_src)
        self._raw_tgt_view.setPlainText(self._raw_tgt)
        self._refresh_tm()
        self._tab_widget().setCurrentIndex(0)
        QMessageBox.information(self, "加载成功", f"已加载 {len(pairs)} 条语料\n请在「原始文档」标签查看")

    def _on_import_tmx(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入 TMX", "", "TMX (*.tmx)")
        if not path: return
        try:
            from ruzh_translator.utils.tmx_parser import parse_tmx
            units = parse_tmx(path); src, tgt = [], []
            for u in units:
                if u.get("source_text"):
                    add_tm_entry(self._session, u["source_text"], u.get("target_text", ""),
                                 u.get("source_lang","ru"), u.get("target_lang","zh-CN"))
                    src.append(u["source_text"]); tgt.append(u.get("target_text", ""))
            self._raw_src = "\n\n".join(src); self._raw_tgt = "\n\n".join(tgt)
            self._raw_src_view.setPlainText(self._raw_src); self._raw_tgt_view.setPlainText(self._raw_tgt)
            self._refresh_tm(); self._tab_widget().setCurrentIndex(0)
            QMessageBox.information(self, "导入成功", f"已导入 {len(units)} 条")
        except Exception as e:
            QMessageBox.critical(self, "导入失败", str(e))

    def _on_import_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入 XLSX", "", "Excel (*.xlsx)")
        if not path: return
        try:
            import openpyxl; wb = openpyxl.load_workbook(path); ws = wb.active
            src, tgt = [], []; count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2: continue
                s = str(row[0]).strip() if row[0] else ""; t = str(row[1]).strip() if len(row)>1 and row[1] else ""
                if not s: continue
                add_tm_entry(self._session, s, t); src.append(s); tgt.append(t); count += 1
            self._raw_src = "\n\n".join(src); self._raw_tgt = "\n\n".join(tgt)
            self._raw_src_view.setPlainText(self._raw_src); self._raw_tgt_view.setPlainText(self._raw_tgt)
            self._refresh_tm(); self._tab_widget().setCurrentIndex(0)
            QMessageBox.information(self, "导入成功", f"已导入 {count} 条")
        except Exception as e:
            QMessageBox.critical(self, "导入失败", str(e))

    # ── TM Browse ────────────────────────────────────────────

    def _refresh_tm(self):
        self._tm_browse.clear()
        entries = self._session.query(TranslationMemoryEntry).order_by(
            TranslationMemoryEntry.usage_count.desc()).limit(200).all()
        for e in entries:
            item = QListWidgetItem(f"{(e.source_text or '')[:60]}\n  → {(e.target_text or '')[:40]}")
            item.setData(Qt.UserRole, e); self._tm_browse.addItem(item)

    def _on_tm_click(self, item):
        e = item.data(Qt.UserRole)
        if e: self._detail_src.setPlainText(e.source_text or ""); self._detail_tgt.setPlainText(e.target_text or "")

    # ── Export ───────────────────────────────────────────────

    def _on_export(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出 KWIC", "kwic.xlsx", "Excel (*.xlsx)")
        if not path: return
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title = "KWIC"
        for c, h in enumerate(["#","左上下文","关键词","右上下文","源语","目标语","标签"],1): ws.cell(1,c,h)
        for r in range(self._kwic_table.rowCount()):
            item = self._kwic_table.item(r, 0)
            if not item: continue
            data = item.data(Qt.UserRole) if hasattr(item, 'data') else {}
            ws.cell(r+2,1,r+1)
            ws.cell(r+2,2,self._kwic_table.item(r,1).text() if self._kwic_table.item(r,1) else "")
            ws.cell(r+2,3,self._kwic_table.item(r,2).text() if self._kwic_table.item(r,2) else "")
            ws.cell(r+2,4,self._kwic_table.item(r,3).text() if self._kwic_table.item(r,3) else "")
            ws.cell(r+2,5,data.get("source_text","") if isinstance(data,dict) else "")
            ws.cell(r+2,6,data.get("target_text","") if isinstance(data,dict) else "")
            tag_w = self._kwic_table.cellWidget(r, 4)
            ws.cell(r+2,7,tag_w.currentText() if isinstance(tag_w, QComboBox) else "")
        wb.save(path)
        QMessageBox.information(self, "导出成功", f"已导出到:\n{path}")
