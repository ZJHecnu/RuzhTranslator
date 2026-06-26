"""Translator Window — OmegaT style with API AI/MT support."""

from pathlib import Path

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QTextEdit, QPushButton, QLabel, QComboBox, QCheckBox, QMessageBox,
    QGroupBox, QTreeWidget, QTreeWidgetItem, QListWidget, QListWidgetItem,
    QProgressBar, QFileDialog,
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont, QColor, QTextCharFormat, QTextCursor

from ruzh_translator.models.base import get_session
from ruzh_translator.models.segment import Segment, AlignmentPair
from ruzh_translator.services.import_service import import_and_segment
from ruzh_translator.services.project_service import create_project, get_project, update_project_status
from ruzh_translator.services.tm_service import fuzzy_match, add_tm_entry, concordance_search
from ruzh_translator.config import IMPORT_FORMATS


class TranslatorWindow(QMainWindow):
    """OmegaT-style translator with 4-panel layout + API translation."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("✏️ 翻译编辑 — Translator")
        self.resize(1500, 950)
        self._session = get_session()
        self._project_id = None
        self._segments = []
        self._current = -1
        self._setup_ui()

    def _setup_ui(self):
        c = QWidget(); self.setCentralWidget(c)
        layout = QVBoxLayout(c); layout.setContentsMargins(8, 8, 8, 8)

        # Top bar
        top = QHBoxLayout()
        top.addWidget(QLabel("<b>✏️ 翻译编辑</b>"))
        self._proj_combo = QComboBox()
        self._proj_combo.addItem("独立翻译（不关联项目）", None)
        self._refresh_projects()
        self._proj_combo.currentIndexChanged.connect(self._on_proj_changed)
        top.addWidget(self._proj_combo, 1)
        top.addWidget(QPushButton("📂 导入源语", clicked=self._on_import_src))
        self._auto_mark = QCheckBox("自动标记已译"); self._auto_mark.setChecked(True)
        top.addWidget(self._auto_mark)
        self._pos_label = QLabel("0 / 0"); self._pos_label.setFixedWidth(70)
        top.addWidget(self._pos_label)
        top.addWidget(QPushButton("◀", clicked=self._on_prev))
        top.addWidget(QPushButton("▶", clicked=self._on_next))
        top.addWidget(QPushButton("💾 保存", clicked=self._on_save))
        top.addWidget(QPushButton("🤖 AI翻译", clicked=self._on_ai_translate))
        top.addWidget(QPushButton("📤 导出", clicked=self._on_export))
        layout.addLayout(top)

        # 4-Panel body
        split = QSplitter(Qt.Horizontal)

        # P1: Navigator
        ng = QGroupBox("句子列表")
        nl = QVBoxLayout(ng)
        self._nav = QTreeWidget(); self._nav.setHeaderLabels(["#", "预览"]); self._nav.setColumnWidth(0, 35); self._nav.setMaximumWidth(250)
        self._nav.currentItemChanged.connect(self._on_nav_click)
        nl.addWidget(self._nav)
        split.addWidget(ng)

        # P2: Source
        sg = QGroupBox("源语（俄语）")
        sl = QVBoxLayout(sg)
        self._src_view = QTextEdit(); self._src_view.setReadOnly(True); self._src_view.setFont(QFont("PingFang SC", 14))
        self._src_view.setStyleSheet("background:#FFFDE7;")
        sl.addWidget(self._src_view)
        split.addWidget(sg)

        # P3: Target
        tg = QGroupBox("译文（中文）")
        tl = QVBoxLayout(tg)
        self._tgt_edit = QTextEdit(); self._tgt_edit.setFont(QFont("PingFang SC", 14))
        self._tgt_edit.setPlaceholderText("在此输入中文译文..."); self._tgt_edit.setStyleSheet("border:2px solid #2196F3;")
        tl.addWidget(self._tgt_edit)
        self._status_lbl = QLabel("状态: ⬜ 未翻译"); self._status_lbl.setStyleSheet("color:#9E9E9E;")
        tl.addWidget(self._status_lbl)
        split.addWidget(tg)

        # P4: TM matches
        tmg = QGroupBox("📖 TM 匹配")
        tml = QVBoxLayout(tmg)
        self._tm_list = QListWidget()
        self._tm_list.setStyleSheet("QListWidget::item{padding:8px;border-bottom:1px solid #EEE;} QListWidget::item:hover{background:#E8F5E9;}")
        self._tm_list.itemDoubleClicked.connect(self._on_tm_insert)
        tml.addWidget(self._tm_list)
        self._tm_info = QLabel(""); self._tm_info.setStyleSheet("color:#9E9E9E;font-size:11px;")
        tml.addWidget(self._tm_info)
        split.addWidget(tmg)

        split.setSizes([220, 420, 420, 220])
        layout.addWidget(split, 1)

        # Bottom: term hints + actions
        bot = QHBoxLayout()
        self._hint_layout = QHBoxLayout(); self._hint_layout.addStretch()
        bot.addLayout(self._hint_layout)
        bot.addStretch()
        for text, fn in [("✅标记已译", lambda: self._set_status("translated")), ("📖查TM", self._on_search_tm), ("🔍术语", self._on_show_terms)]:
            bot.addWidget(QPushButton(text, clicked=fn))
        layout.addLayout(bot)

    def _refresh_projects(self):
        from ruzh_translator.services.project_service import list_projects
        for p in list_projects(self._session):
            self._proj_combo.addItem(p.name, p.id)

    def _on_proj_changed(self):
        pid = self._proj_combo.currentData()
        if pid:
            self._project_id = pid
            pairs = self._session.query(AlignmentPair).filter(AlignmentPair.project_id == pid).order_by(AlignmentPair.paragraph_index, AlignmentPair.pair_index).all()
            if pairs:
                self._segments = pairs
            else:
                self._segments = self._session.query(Segment).filter(Segment.project_id == pid).order_by(Segment.paragraph_index, Segment.segment_index).all()
            self._build_nav()
            if self._segments: self._navigate(0)

    def _on_import_src(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入源语文档", "", ";;".join(f"{d} (*{e})" for e, d in IMPORT_FORMATS.items()))
        if not path: return
        if not self._project_id:
            p = create_project(self._session, "翻译项目"); self._project_id = p.id; self._proj_combo.addItem(p.name, p.id)
        result = import_and_segment(path, self._session, self._project_id)
        self._segments = result["segments"]
        self._build_nav()
        if self._segments: self._navigate(0)
        update_project_status(self._session, self._project_id, "translation")

    def _build_nav(self):
        self._nav.clear(); cur = -1
        icons = {"untranslated":"⬜","draft":"📝","translated":"✅","reviewed":"✓","approved":"⭐"}
        for i, item in enumerate(self._segments):
            pi = getattr(item, 'paragraph_index', 0) or 0
            if pi != cur:
                cur = pi; pitem = QTreeWidgetItem(self._nav); pitem.setText(0, f"段{cur+1}"); pitem.setData(0, Qt.UserRole, -1); pitem.setExpanded(True)
            src = self._src(item)
            ti = QTreeWidgetItem(pitem); ti.setText(0, str(i+1))
            st = getattr(item, 'status', 'untranslated') or 'untranslated'
            ti.setText(1, f"{icons.get(st,'⬜')} {(src or '')[:25]}..."); ti.setData(0, Qt.UserRole, i)

    def _src(self, item):
        return (item.source_text or "") if hasattr(item, 'source_text') else ""
    def _tgt(self, item):
        return (item.target_text or "") if hasattr(item, 'target_text') else ""
    def _set_tgt(self, item, v):
        if hasattr(item, 'target_text'): item.target_text = v

    def _navigate(self, idx):
        if idx < 0 or idx >= len(self._segments): return
        if self._current >= 0 and self._current != idx:
            self._save_current()
            if self._auto_mark.isChecked() and self._tgt(self._segments[self._current]).strip():
                self._segments[self._current].status = "translated"; self._session.commit()
        self._current = idx; item = self._segments[idx]
        self._src_view.setPlainText(self._src(item)); self._tgt_edit.setPlainText(self._tgt(item))
        self._pos_label.setText(f"{idx+1} / {len(self._segments)}")
        st = getattr(item, 'status', 'untranslated') or 'untranslated'
        sm = {"untranslated":"⬜ 未翻译","draft":"📝 草稿","translated":"✅ 已翻译","reviewed":"✓ 已审校","approved":"⭐ 已批准"}
        self._status_lbl.setText(f"状态: {sm.get(st, st)}")
        self._refresh_tm(); self._refresh_hints()
        self._tgt_edit.setFocus()

    def _on_nav_click(self, cur, _):
        if cur:
            idx = cur.data(0, Qt.UserRole)
            if idx is not None and idx >= 0: self._navigate(idx)

    def _on_prev(self): self._navigate(self._current - 1) if self._current > 0 else None
    def _on_next(self):
        if self._current < len(self._segments) - 1: self._navigate(self._current + 1)

    def _save_current(self):
        if self._current < 0: return
        item = self._segments[self._current]; nt = self._tgt_edit.toPlainText().strip()
        if nt and nt != self._tgt(item):
            self._set_tgt(item, nt); item.status = item.status or "draft"; item.is_manually_corrected = True
            self._session.commit()
            try: add_tm_entry(self._session, self._src(item), nt, project_id=self._project_id)
            except: pass

    def _on_save(self):
        self._save_current(); self._update_progress()

    def _set_status(self, st):
        self._save_current()
        if self._current >= 0: self._segments[self._current].status = st; self._session.commit(); self._build_nav(); self._navigate(self._current)

    def _update_progress(self):
        t = len(self._segments); tr = sum(1 for s in self._segments if getattr(s,'status','') in ('translated','reviewed','approved'))
        self.setWindowTitle(f"✏️ 翻译编辑 — Translator  [{tr}/{t}]")

    def _refresh_tm(self):
        self._tm_list.clear()
        if self._current < 0: return
        src = self._src(self._segments[self._current])
        if not src: return
        try:
            results = fuzzy_match(self._session, src, threshold=0.55, limit=8)
            for r in results:
                pct = int(r["score"]*100); bar = "█"*(pct//10)+"░"*(10-pct//10)
                item = QListWidgetItem(f"{bar} {pct}%\n{r['entry'].target_text[:80]}")
                item.setData(Qt.UserRole, r["entry"].target_text); self._tm_list.addItem(item)
            self._tm_info.setText(f"找到 {len(results)} 条匹配" if results else "未找到匹配")
        except: pass

    def _on_tm_insert(self, item):
        t = item.data(Qt.UserRole)
        if t: self._tgt_edit.setPlainText(t)

    def _refresh_hints(self):
        while self._hint_layout.count() > 1:
            w = self._hint_layout.takeAt(0).widget()
            if w: w.deleteLater()
        if self._current < 0 or not self._project_id: return
        src = self._src(self._segments[self._current])
        if not src: return
        try:
            from ruzh_translator.services.term_hint_service import get_term_hints
            for h in get_term_hints(self._session, src, self._project_id)[:6]:
                btn = QPushButton(f"{h['source_term']} → {h['target_term']}")
                btn.setStyleSheet("background:#E3F2FD;color:#1565C0;border:1px solid #90CAF9;border-radius:12px;padding:4px 10px;font-size:12px;")
                btn.clicked.connect(lambda checked, t=h['target_term']: self._tgt_edit.textCursor().insertText(t))
                self._hint_layout.insertWidget(self._hint_layout.count()-1, btn)
        except: pass

    def _on_ai_translate(self):
        """Translate current segment using AI API."""
        if self._current < 0: return
        src = self._src(self._segments[self._current])
        if not src: return
        from ruzh_translator.ui.settings_dialog import get_ai_config
        cfg = get_ai_config()
        if not cfg["enabled"] or not cfg["key"]:
            QMessageBox.warning(self, "提示", "请先在偏好设置中配置 AI API"); return
        try:
            from ruzh_translator.services.term_extraction_service import translate_terms_with_ai
            import urllib.request, json
            req = urllib.request.Request(f"{cfg['url'].rstrip('/')}/chat/completions",
                data=json.dumps({"model":cfg["model"],"messages":[
                    {"role":"system","content":"你是专业俄汉翻译。只输出中文译文，不要解释。"},
                    {"role":"user","content":f"将以下俄语翻译为中文:\n{src}"}
                ],"temperature":0.3}).encode(),
                headers={"Authorization":f"Bearer {cfg['key']}","Content-Type":"application/json"})
            resp = json.loads(urllib.request.urlopen(req, timeout=30).read())
            self._tgt_edit.setPlainText(resp["choices"][0]["message"]["content"].strip())
        except Exception as e:
            QMessageBox.critical(self, "AI 翻译失败", str(e))

    def _on_search_tm(self):
        if self._current < 0: return
        src = self._src(self._segments[self._current])
        if not src: return
        results = fuzzy_match(self._session, src, threshold=0.6, limit=5)
        msg = "翻译记忆匹配:\n\n" + "\n\n".join(f"[{r['score']:.0%}] {r['entry'].target_text[:120]}" for r in results) if results else "未找到匹配"
        QMessageBox.information(self, "TM 查询", msg)

    def _on_show_terms(self):
        if not self._project_id: return
        from ruzh_translator.models.glossary import GlossaryEntry
        entries = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id).limit(100).all()
        msg = f"术语 ({len(entries)}):\n\n" + "\n".join(f"• {e.source_term} → {e.target_term}" for e in entries[:40]) if entries else "暂无术语"
        QMessageBox.information(self, "术语", msg)

    def _on_export(self):
        if not self._segments: return
        path, _ = QFileDialog.getSaveFileName(self, "导出译文", "translation.xlsx", "Excel (*.xlsx);;TMX (*.tmx)")
        if not path: return
        if path.endswith('.tmx'):
            from ruzh_translator.utils.tmx_parser import export_tmx
            pairs = [{"source_text":self._src(s),"target_text":self._tgt(s)} for s in self._segments if self._tgt(s).strip()]
            export_tmx(pairs, path)
        else:
            import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title = "翻译结果"
            for c, h in enumerate(["序号","源语","译文","状态"],1): ws.cell(1,c,h)
            for i, s in enumerate(self._segments):
                ws.cell(i+2,1,i+1); ws.cell(i+2,2,self._src(s)); ws.cell(i+2,3,self._tgt(s)); ws.cell(i+2,4,getattr(s,'status',''))
            wb.save(path)
        QMessageBox.information(self, "导出成功", f"已导出到:\n{path}")
