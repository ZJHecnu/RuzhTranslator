"""Project creation wizard: 3-step dialog.

Step 1: Basic info (name, language pair)
Step 2: Workflow selection (align / translate / review / glossary / full)
Step 3: File import (dynamic based on workflow choice)
"""

from pathlib import Path

from PySide6.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QComboBox, QTextEdit, QPushButton,
    QRadioButton, QButtonGroup, QGroupBox, QFileDialog,
    QStackedWidget, QMessageBox, QProgressBar,
)
from PySide6.QtCore import Qt, Signal

from ruzh_translator.config import SOURCE_LANGS, TARGET_LANGS, IMPORT_FORMATS
from ruzh_translator.models.base import get_session
from ruzh_translator.services.project_service import create_project, create_project_folder
from ruzh_translator.services.import_service import import_and_segment, import_file


WORKFLOWS = [
    ("full",       "🚀 完整流程",        "从导入到翻译到审校到导出（推荐）"),
    ("align",      "🔗 仅对齐",          "我有源语和目标语文件，需要句子对齐"),
    ("translate",  "✏️ 仅翻译",          "我有源语文件，需要逐句翻译"),
    ("review",     "✓  仅审校",          "我有已翻译的文件（xlsx/tmx），需要审校"),
    ("glossary",   "📖 仅术语",          "我有源语文件，需要提取和管理术语"),
]


class ProjectWizard(QDialog):
    """Three-step project creation wizard."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建项目")
        self.setMinimumSize(650, 520)
        self._session = get_session()

        # Results
        self.project_id: str | None = None
        self.project_name: str = ""
        self.source_lang: str = "ru"
        self.target_lang: str = "zh-CN"
        self.description: str = ""
        self.workflow: str = "full"
        self.source_path: str = ""
        self.target_path: str = ""
        self.glossary_path: str = ""
        self.tm_path: str = ""
        self.review_path: str = ""

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(0)

        # ── Step indicator ──
        self._step_label = QLabel("<b>步骤 1/3: 基本信息</b>")
        self._step_label.setStyleSheet("padding: 12px 16px; color: #1565C0; font-size: 15px;")
        layout.addWidget(self._step_label)

        # ── Stacked pages ──
        self._stack = QStackedWidget()
        self._stack.addWidget(self._make_step1())
        self._stack.addWidget(self._make_step2())
        self._stack.addWidget(self._make_step3())
        layout.addWidget(self._stack, 1)

        # ── Navigation buttons ──
        nav = QHBoxLayout()
        nav.setContentsMargins(16, 8, 16, 12)

        self._back_btn = QPushButton("← 上一步")
        self._back_btn.clicked.connect(self._on_back)
        self._back_btn.setVisible(False)
        nav.addWidget(self._back_btn)

        nav.addStretch()

        self._next_btn = QPushButton("下一步 →")
        self._next_btn.setStyleSheet(
            "QPushButton { background: #1565C0; color: white; padding: 8px 24px; "
            "border-radius: 6px; font-size: 14px; } "
            "QPushButton:hover { background: #1976D2; }"
        )
        self._next_btn.clicked.connect(self._on_next)
        nav.addWidget(self._next_btn)

        layout.addLayout(nav)

    # ── Step 1: Basic Info ────────────────────────────────────

    def _make_step1(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(24, 16, 24, 16)

        form = QFormLayout()
        form.setSpacing(14)

        self._name_edit = QLineEdit()
        self._name_edit.setPlaceholderText("输入项目名称，如：邓小平文选第三章")
        form.addRow("项目名称:", self._name_edit)

        src_layout = QHBoxLayout()
        self._src_combo = QComboBox()
        for code, name in SOURCE_LANGS.items():
            self._src_combo.addItem(name, code)
        src_layout.addWidget(self._src_combo, 1)
        tgt_label = QLabel("  →  ")
        src_layout.addWidget(tgt_label)
        self._tgt_combo = QComboBox()
        for code, name in TARGET_LANGS.items():
            self._tgt_combo.addItem(name, code)
        self._tgt_combo.setCurrentIndex(1)  # zh-CN
        src_layout.addWidget(self._tgt_combo, 1)
        form.addRow("语言对:", src_layout)

        self._desc_edit = QTextEdit()
        self._desc_edit.setPlaceholderText("项目描述（可选）")
        self._desc_edit.setMaximumHeight(60)
        form.addRow("描述:", self._desc_edit)

        layout.addLayout(form)
        layout.addStretch()
        return page

    # ── Step 2: Workflow Selection ────────────────────────────

    def _make_step2(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(24, 16, 24, 16)

        layout.addWidget(QLabel("<b>你想用这个项目做什么？</b>"))
        layout.addSpacing(8)

        self._workflow_group = QButtonGroup(self)
        self._workflow_radios = {}

        for i, (key, title, desc) in enumerate(WORKFLOWS):
            radio = QRadioButton(f"{title}  —  {desc}")
            radio.setStyleSheet(
                "QRadioButton { padding: 10px 14px; font-size: 13px; } "
                "QRadioButton::indicator { width: 18px; height: 18px; }"
            )
            self._workflow_group.addButton(radio, i)
            self._workflow_radios[key] = radio
            layout.addWidget(radio)

        # Default: full workflow
        self._workflow_radios["full"].setChecked(True)

        layout.addStretch()
        return page

    # ── Step 3: File Import ───────────────────────────────────

    def _make_step3(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(24, 16, 24, 16)

        self._step3_label = QLabel("<b>导入文件</b>")
        layout.addWidget(self._step3_label)

        self._step3_desc = QLabel("根据选择的工作流，导入相应文件")
        self._step3_desc.setStyleSheet("color: #9E9E9E;")
        layout.addWidget(self._step3_desc)

        layout.addSpacing(8)

        # File rows (visibility toggled based on workflow)
        self._file_rows = {}

        # Source file (always shown except review)
        self._file_rows["source"] = self._make_file_row("📂 源语文档（必选）:", self._on_browse_source)
        layout.addLayout(self._file_rows["source"])

        # Target file (align + full)
        self._file_rows["target"] = self._make_file_row("📂 目标语文档（可选）:", self._on_browse_target)
        layout.addLayout(self._file_rows["target"])

        # Glossary file (translate + full + glossary)
        self._file_rows["glossary"] = self._make_file_row("📂 术语表（可选，.xlsx）:", self._on_browse_glossary)
        layout.addLayout(self._file_rows["glossary"])

        # TM file (translate + full)
        self._file_rows["tm"] = self._make_file_row("📂 翻译记忆（可选，.tmx）:", self._on_browse_tm)
        layout.addLayout(self._file_rows["tm"])

        # Review file (review only)
        self._file_rows["review"] = self._make_file_row("📂 翻译文件（.xlsx / .tmx）:", self._on_browse_review)
        layout.addLayout(self._file_rows["review"])

        layout.addStretch()

        # Progress
        self._progress = QProgressBar()
        self._progress.setVisible(False)
        layout.addWidget(self._progress)

        return page

    def _make_file_row(self, label_text: str, browse_handler) -> QHBoxLayout:
        """Create a file selection row with label, status, and browse button."""
        row = QHBoxLayout()
        label = QLabel(label_text)
        label.setFixedWidth(250)
        row.addWidget(label)

        status = QLabel("未选择")
        status.setStyleSheet("color: #9E9E9E;")
        status.setObjectName("status")
        row.addWidget(status, 1)

        btn = QPushButton("选择文件...")
        btn.clicked.connect(browse_handler)
        btn.setObjectName("browse")
        row.addWidget(btn)

        return row

    # ── File Browsers ────────────────────────────────────────

    def _browse(self, key: str, setter: str, filters: str = None):
        if not filters:
            filters = ";;".join(f"{d} (*{e})" for e, d in IMPORT_FORMATS.items())
        path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", filters)
        if path:
            setattr(self, setter, path)
            # Update status label
            row = self._file_rows[key]
            for i in range(row.count()):
                w = row.itemAt(i).widget()
                if w and w.objectName() == "status":
                    w.setText(f"✓ {Path(path).name}")
                    w.setStyleSheet("color: #4CAF50; font-weight: bold;")
                    break

    def _on_browse_source(self):
        self._browse("source", "source_path")
    def _on_browse_target(self):
        self._browse("target", "target_path")
    def _on_browse_glossary(self):
        self._browse("glossary", "glossary_path", "Excel 文件 (*.xlsx)")
    def _on_browse_tm(self):
        self._browse("tm", "tm_path", "TMX 文件 (*.tmx)")
    def _on_browse_review(self):
        self._browse("review", "review_path", "Excel/TMX 文件 (*.xlsx *.tmx)")

    # ── Navigation ───────────────────────────────────────────

    def _on_back(self):
        idx = self._stack.currentIndex()
        if idx > 0:
            self._stack.setCurrentIndex(idx - 1)
        self._update_nav()

    def _on_next(self):
        idx = self._stack.currentIndex()

        if idx == 0:
            # Validate step 1
            if not self._name_edit.text().strip():
                QMessageBox.warning(self, "提示", "请输入项目名称")
                return
            self.project_name = self._name_edit.text().strip()
            self.source_lang = self._src_combo.currentData()
            self.target_lang = self._tgt_combo.currentData()
            self.description = self._desc_edit.toPlainText().strip()

        elif idx == 1:
            # Get selected workflow
            for key, radio in self._workflow_radios.items():
                if radio.isChecked():
                    self.workflow = key
                    break
            # Configure step 3 UI
            self._configure_step3(self.workflow)

        elif idx == 2:
            # Validate and create project
            if not self._validate_step3():
                return
            self._create_project()
            return

        self._stack.setCurrentIndex(idx + 1)
        self._update_nav()

    def _update_nav(self):
        idx = self._stack.currentIndex()
        self._step_label.setText(f"<b>步骤 {idx + 1}/3: {['基本信息', '工作流选择', '导入文件'][idx]}</b>")
        self._back_btn.setVisible(idx > 0)
        self._next_btn.setText("创建项目" if idx == 2 else "下一步 →")

    # ── Step 3 Configuration ─────────────────────────────────

    def _configure_step3(self, workflow: str):
        """Show/hide file rows based on workflow choice."""
        visibility = {
            "full":      {"source": True, "target": True, "glossary": True, "tm": True, "review": False},
            "align":     {"source": True, "target": True, "glossary": False, "tm": False, "review": False},
            "translate": {"source": True, "target": False, "glossary": True, "tm": True, "review": False},
            "review":    {"source": False, "target": False, "glossary": False, "tm": False, "review": True},
            "glossary":  {"source": True, "target": False, "glossary": False, "tm": False, "review": False},
        }

        vis = visibility.get(workflow, visibility["full"])
        for key, row in self._file_rows.items():
            for i in range(row.count()):
                w = row.itemAt(i).widget()
                if w:
                    w.setVisible(vis.get(key, False))

    def _validate_step3(self) -> bool:
        """Validate that required files are selected."""
        wf = self.workflow
        if wf in ("full", "translate", "glossary") and not self.source_path:
            QMessageBox.warning(self, "提示", "请选择源语文档")
            return False
        if wf in ("align",) and (not self.source_path or not self.target_path):
            QMessageBox.warning(self, "提示", "仅对齐模式需要源语和目标语文档")
            return False
        if wf in ("review",) and not self.review_path:
            QMessageBox.warning(self, "提示", "请选择需要审校的翻译文件")
            return False
        return True

    # ── Project Creation ─────────────────────────────────────

    def _create_project(self):
        """Create the project, import files, and return results."""
        self._progress.setVisible(True)
        self._progress.setRange(0, 0)

        try:
            # Step A: Create project in DB + folder
            proj = create_project(
                self._session,
                self.project_name,
                self.source_lang,
                self.target_lang,
                self.description,
            )
            self.project_id = proj.id

            # Step B: Import source file (unless review mode)
            if self.source_path:
                result = import_and_segment(
                    self.source_path, self._session, proj.id,
                    language=self.source_lang,
                )
                proj.status = (
                    "alignment" if self.workflow == "align" else "translation"
                )
                self._session.commit()

            # Step C: Import target file (align mode)
            if self.target_path and self.workflow in ("full", "align"):
                import_and_segment(
                    self.target_path, self._session, proj.id,
                    language=self.target_lang,
                )

            # Step D: Import review file (review mode)
            if self.review_path:
                self._import_review_file()

            # Step E: Import glossary if provided
            if self.glossary_path:
                self._import_glossary()

            # Step F: Import TM if provided
            if self.tm_path:
                self._import_tm()

            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "创建失败", str(e))
        finally:
            self._progress.setVisible(False)

    def _import_review_file(self):
        """Parse xlsx/tmx review file into Segments."""
        suffix = Path(self.review_path).suffix.lower()

        if suffix == ".tmx":
            from ruzh_translator.utils.tmx_parser import parse_tmx
            units = parse_tmx(self.review_path)
            from ruzh_translator.models.segment import Segment
            for i, u in enumerate(units):
                seg = Segment(
                    project_id=self.project_id,
                    paragraph_index=0,
                    segment_index=i,
                    source_text=u.get("source_text", ""),
                    target_text=u.get("target_text", ""),
                    status="translated",
                )
                self._session.add(seg)

        elif suffix in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(self.review_path)
            ws = wb.active
            from ruzh_translator.models.segment import Segment
            seg_idx = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row or len(row) < 2:
                    continue
                src = str(row[0]).strip() if row[0] else ""
                tgt = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if not src:
                    continue
                seg = Segment(
                    project_id=self.project_id,
                    paragraph_index=0,
                    segment_index=seg_idx,
                    source_text=src,
                    target_text=tgt,
                    status="translated",
                )
                self._session.add(seg)
                seg_idx += 1

        self._session.commit()

    def _import_glossary(self):
        """Import glossary entries from xlsx."""
        import openpyxl
        from ruzh_translator.models.glossary import GlossaryEntry
        wb = openpyxl.load_workbook(self.glossary_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            src = str(row[0]).strip() if row[0] else ""
            tgt = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not src:
                continue
            entry = GlossaryEntry(
                project_id=self.project_id,
                source_term=src,
                target_term=tgt,
                is_approved=1,
            )
            self._session.add(entry)
        self._session.commit()

    def _import_tm(self):
        """Import TM entries from tmx."""
        from ruzh_translator.utils.tmx_parser import import_tmx_to_db
        import_tmx_to_db(self.tm_path, self._session, project_id=self.project_id)
