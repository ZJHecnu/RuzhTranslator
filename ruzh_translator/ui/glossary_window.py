"""Glossary Window — term management + AI auto-translation."""

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QLabel,
    QLineEdit, QFileDialog, QMessageBox, QProgressBar, QComboBox,
)
from PySide6.QtCore import Qt

from ruzh_translator.models.base import get_session
from ruzh_translator.models.glossary import GlossaryEntry
from ruzh_translator.models.segment import Segment
from ruzh_translator.services.term_extraction_service import extract_terms_from_segments, translate_terms_with_ai
from ruzh_translator.services.project_service import list_projects
from ruzh_translator.config import IMPORT_FORMATS


class GlossaryWindow(QMainWindow):
    """Glossary manager with auto-extraction + AI translation."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("📖 术语管理 — Glossary")
        self.resize(1000, 750)
        self._session = get_session()
        self._project_id = None
        self._setup_ui()

    def _setup_ui(self):
        c = QWidget(); self.setCentralWidget(c)
        layout = QVBoxLayout(c); layout.setContentsMargins(8, 8, 8, 8)

        # Toolbar
        tools = QHBoxLayout()
        tools.addWidget(QLabel("<b>📖 术语管理</b>"))

        self._proj_combo = QComboBox()
        self._proj_combo.addItem("独立术语库（不关联项目）", None)
        for p in list_projects(self._session):
            self._proj_combo.addItem(p.name, p.id)
        self._proj_combo.currentIndexChanged.connect(self._on_proj_changed)
        tools.addWidget(self._proj_combo, 1)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("搜索术语...")
        self._search_edit.textChanged.connect(self._refresh)
        tools.addWidget(self._search_edit)

        tools.addWidget(QPushButton("🔍 提取术语", clicked=self._on_extract))
        tools.addWidget(QPushButton("🤖 AI翻译术语", clicked=self._on_ai_translate))
        tools.addWidget(QPushButton("➕ 添加", clicked=self._on_add))
        tools.addWidget(QPushButton("📥 导入", clicked=self._on_import))
        tools.addWidget(QPushButton("📤 导出", clicked=self._on_export))
        layout.addLayout(tools)

        self._progress = QProgressBar(); self._progress.setVisible(False)
        layout.addWidget(self._progress)

        # Table
        self._table = QTableWidget()
        self._table.setColumnCount(5)
        self._table.setHorizontalHeaderLabels(["源语术语", "目标语翻译", "领域", "状态", "备注"])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self._table, 1)

        # Bottom actions
        bot = QHBoxLayout()
        bot.addWidget(QPushButton("✏️ 编辑选中", clicked=self._on_edit))
        bot.addWidget(QPushButton("✓ 确认选中", clicked=lambda: self._on_approve(True)))
        bot.addWidget(QPushButton("🗑 删除选中", clicked=self._on_delete))
        bot.addStretch()
        self._count_label = QLabel("共 0 条术语")
        self._count_label.setStyleSheet("color:#9E9E9E;")
        bot.addWidget(self._count_label)
        layout.addLayout(bot)

        self._refresh()

    def _on_proj_changed(self):
        self._project_id = self._proj_combo.currentData()
        self._refresh()

    def _refresh(self):
        if not self._project_id:
            self._table.setRowCount(0); return
        q = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id)
        s = self._search_edit.text().strip()
        if s: q = q.filter(GlossaryEntry.source_term.contains(s))
        entries = q.order_by(GlossaryEntry.source_term).all()
        self._table.setRowCount(len(entries))
        for r, e in enumerate(entries):
            self._table.setItem(r, 0, QTableWidgetItem(e.source_term))
            self._table.setItem(r, 1, QTableWidgetItem(e.target_term))
            self._table.setItem(r, 2, QTableWidgetItem(e.domain or ""))
            self._table.setItem(r, 3, QTableWidgetItem("✓已确认" if e.is_approved else "待确认"))
            self._table.setItem(r, 4, QTableWidgetItem((e.notes or "")[:50]))
        self._count_label.setText(f"共 {len(entries)} 条术语")

    def _on_extract(self):
        if not self._project_id: QMessageBox.warning(self, "提示", "请先选择项目"); return
        segs = self._session.query(Segment).filter(Segment.project_id == self._project_id).all()
        if not segs: QMessageBox.warning(self, "提示", "该项目没有句子数据，请先在翻译模块中导入源语文档"); return
        self._progress.setVisible(True); self._progress.setRange(0, 0)
        try:
            terms = extract_terms_from_segments(segs, language="ru", top_n=80)
            added = 0
            for td in terms:
                existing = self._session.query(GlossaryEntry).filter(
                    GlossaryEntry.project_id == self._project_id,
                    GlossaryEntry.source_term == td["term"]
                ).first()
                if not existing:
                    self._session.add(GlossaryEntry(project_id=self._project_id, source_term=td["term"], target_term="", is_approved=0, notes=f"自动提取 (频率:{td.get('frequency',0)})"))
                    added += 1
            self._session.commit(); self._refresh()
            QMessageBox.information(self, "提取完成", f"提取到 {len(terms)} 个候选术语，新增 {added} 个")
        except Exception as e:
            QMessageBox.critical(self, "提取失败", str(e))
        finally:
            self._progress.setVisible(False)

    def _on_ai_translate(self):
        """Use AI to translate unconfirmed terms."""
        from ruzh_translator.ui.settings_dialog import get_ai_config
        cfg = get_ai_config()
        if not cfg["enabled"] or not cfg["key"]:
            QMessageBox.warning(self, "提示", "请先在偏好设置中配置 AI API\n(菜单栏: 设置 → 偏好设置)"); return

        entries = self._session.query(GlossaryEntry).filter(
            GlossaryEntry.project_id == self._project_id,
            GlossaryEntry.is_approved == 0
        ).all()
        if not entries: QMessageBox.information(self, "提示", "没有待确认的术语"); return

        terms = [e.source_term for e in entries if not e.target_term]
        if not terms: QMessageBox.information(self, "提示", "所有待确认术语都已有翻译"); return

        self._progress.setVisible(True); self._progress.setRange(0, 0)
        try:
            translations = translate_terms_with_ai(terms)
            count = 0
            for e in entries:
                if e.source_term in translations:
                    e.target_term = translations[e.source_term]; e.is_approved = 1; count += 1
            self._session.commit(); self._refresh()
            QMessageBox.information(self, "AI 翻译完成", f"已翻译 {count} 个术语")
        except Exception as e:
            QMessageBox.critical(self, "AI 翻译失败", str(e))
        finally:
            self._progress.setVisible(False)

    def _on_add(self):
        if not self._project_id: return
        from PySide6.QtWidgets import QDialog, QFormLayout, QDialogButtonBox, QTextEdit
        dlg = QDialog(self); dlg.setWindowTitle("添加术语"); dlg.setMinimumWidth(400)
        l = QFormLayout(dlg)
        src_e = QLineEdit(); tgt_e = QLineEdit(); dom_e = QLineEdit()
        l.addRow("源语术语:", src_e); l.addRow("目标语翻译:", tgt_e); l.addRow("领域:", dom_e)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(lambda: (
            setattr(dlg, 'result', (src_e.text().strip(), tgt_e.text().strip(), dom_e.text().strip())),
            dlg.accept()
        ) if src_e.text().strip() else QMessageBox.warning(dlg, "提示", "请输入源语术语"))
        btns.rejected.connect(dlg.reject)
        l.addRow(btns)
        if dlg.exec():
            src, tgt, dom = dlg.result
            self._session.add(GlossaryEntry(project_id=self._project_id, source_term=src, target_term=tgt, domain=dom, is_approved=1))
            self._session.commit(); self._refresh()

    def _on_edit(self):
        r = self._table.currentRow()
        if r < 0: return
        src = self._table.item(r, 0).text()
        e = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id, GlossaryEntry.source_term == src).first()
        if not e: return
        from PySide6.QtWidgets import QDialog, QFormLayout, QDialogButtonBox
        dlg = QDialog(self); dlg.setWindowTitle("编辑术语")
        fl = QFormLayout(dlg)
        tgt_e = QLineEdit(e.target_term); dom_e = QLineEdit(e.domain or "")
        fl.addRow("源语:", QLabel(e.source_term))
        fl.addRow("目标语:", tgt_e); fl.addRow("领域:", dom_e)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        fl.addRow(btns)
        if dlg.exec():
            e.target_term = tgt_e.text().strip(); e.domain = dom_e.text().strip(); e.is_approved = 1
            self._session.commit(); self._refresh()

    def _on_approve(self, approved: bool):
        r = self._table.currentRow()
        if r < 0: return
        src = self._table.item(r, 0).text()
        e = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id, GlossaryEntry.source_term == src).first()
        if e: e.is_approved = 1 if approved else 0; self._session.commit(); self._refresh()

    def _on_delete(self):
        r = self._table.currentRow()
        if r < 0: return
        src = self._table.item(r, 0).text()
        if QMessageBox.question(self, "确认", f"删除术语「{src}」？") == QMessageBox.Yes:
            e = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id, GlossaryEntry.source_term == src).first()
            if e: self._session.delete(e); self._session.commit(); self._refresh()

    def _on_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入术语表", "", "Excel (*.xlsx)")
        if not path: return
        if not self._project_id:
            from ruzh_translator.services.project_service import create_project
            p = create_project(self._session, "术语项目"); self._project_id = p.id; self._proj_combo.addItem(p.name, p.id)
        import openpyxl; wb = openpyxl.load_workbook(path); ws = wb.active; count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 1: continue
            src = str(row[0]).strip() if row[0] else ""; tgt = str(row[1]).strip() if len(row)>1 and row[1] else ""
            if not src: continue
            existing = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id, GlossaryEntry.source_term == src).first()
            if existing:
                if tgt: existing.target_term = tgt
            else:
                self._session.add(GlossaryEntry(project_id=self._project_id, source_term=src, target_term=tgt, is_approved=1))
            count += 1
        self._session.commit(); self._refresh()
        QMessageBox.information(self, "导入成功", f"已导入 {count} 条术语")

    def _on_export(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出术语表", "glossary.xlsx", "Excel (*.xlsx)")
        if not path: return
        entries = self._session.query(GlossaryEntry).filter(GlossaryEntry.project_id == self._project_id).all()
        import openpyxl; wb = openpyxl.Workbook(); ws = wb.active; ws.title = "术语表"
        for c, h in enumerate(["源语术语","目标语翻译","领域","状态","备注"],1): ws.cell(1,c,h)
        for i, e in enumerate(entries):
            ws.cell(i+2,1,e.source_term); ws.cell(i+2,2,e.target_term); ws.cell(i+2,3,e.domain or "")
            ws.cell(i+2,4,"已确认" if e.is_approved else "待确认"); ws.cell(i+2,5,e.notes or "")
        wb.save(path)
        QMessageBox.information(self, "导出成功", f"已导出 {len(entries)} 条术语到:\n{path}")
